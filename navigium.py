import requests
from bs4 import BeautifulSoup
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from io import BytesIO
from export_to_Brainyoo import save_by2
import concurrent.futures
import locale
from termcolor import colored
from openpyxl.styles import Font

locale.setlocale(locale.LC_ALL, 'de_DE.UTF-8')
locale.setlocale(locale.LC_ALL, 'C')
def dice(scriptio, verbum, iudicatum="good"):
    if iudicatum == "good":
        color = "green"
    elif iudicatum == "bad":
        color = "red"
    print(colored(scriptio, color) + ": " + colored(verbum, 'yellow'))

def request2navigium(words=["currere"]):
    output = []
    format = {
        "grundform": str,
        "bedeutungen": list,
        "flexion": list,
        "wortart": str,
        "deponens": bool,
        "klassenlabel": str
        }
    for word in words:
        url = f"https://www.navigium.de/suchfunktion/_search?q={word}"
        try:
            json_file = requests.get(url).content
            data = json.loads(json_file)
        except:
            dice("Error: Wort konnte auf Navigium nicht gefunden werden", word, "bad")
            continue

        try:
            answer = data[0]["searchItems"][0]
        except:
            dice("Error: Wort konnte auf Navigium nicht gefunden werden bzw. grundlegende Informationen nicht vorhanden", word, "bad")
            continue

        formated_answer = format.copy()
        try:
            formated_answer["grundform"] = answer["lemma"]
            formated_answer["bedeutungen"] = answer["bedeutungsgruppe"][0]["bedeutungJoined"]
            formated_answer["flexion"] = answer["flexion"]
            formated_answer["wortart"] = answer["wortart"]
            formated_answer["deponens"] = answer["deponens"]
            try:
                formated_answer["klassenlabel"] = answer["klassenlabel"]
            except:
                formated_answer["klassenlabel"] = "Keine Angabe"
        except:
            dice("Error: Informationen konnten nicht extrahiert werden", word, "bad")
            continue
        output.append(formated_answer)
        dice("Erfolgreich: Alle Informationen konnten erfolgreich extrahiert werden", word, "good")
    return output

def threaded_function(words, function=request2navigium, anzahl_threads=20, wordtype=None):
    # Erstellen von Blöcken für jeden Thread
    blockgröße = len(words) // anzahl_threads
    rest = len(words) % anzahl_threads

    # Verteilen der Namen auf die Threads
    threads_aufgaben = []
    start_index = 0

    for i in range(anzahl_threads):
        zusätzliche_namen = 1 if i < rest else 0
        end_index = start_index + blockgröße + zusätzliche_namen
        threads_aufgaben.append(words[start_index:end_index])
        start_index = end_index

    # Liste für alle Ergebnisse
    alle_ergebnisse = []

    # Verwenden von Threads, um die Namen zu verarbeiten
    with concurrent.futures.ThreadPoolExecutor(max_workers=anzahl_threads) as executor:
        # Future-Objekte sammeln
        if wordtype == None:
            future_to_result = {executor.submit(function, word): word 
                            for word in threads_aufgaben}
        else:
            future_to_result = {executor.submit(function, word, wordtype): word 
                            for word in threads_aufgaben}
        
        # Ergebnisse sammeln
        for future in concurrent.futures.as_completed(future_to_result):
            try:
                ergebnis = future.result()
                alle_ergebnisse.extend(ergebnis)
            except Exception as e:
                print(f"Ein Fehler ist aufgetreten: {e}")

    # Hier hast du alle Ergebnisse in der Variable 'alle_ergebnisse'
    return alle_ergebnisse

def sort_by_wordtype(input):
    sorted = {
        "Nomen": [],
        "Verben": [],
        "Adjektive": [],
        "Adverbien": [],
        "Pronomen": [],
        "Konjunktionen": [],
        "Präpositionen": [],
        "Subjunktionen": [],
        "Unbekannt": []
    }

    for word in input:
        if word["wortart"].lower() == "SUBST".lower():
            sorted["Nomen"].append(word)
        elif word["wortart"].lower() == "VERB".lower():
            sorted["Verben"].append(word)
        elif word["wortart"].lower()  == "ADJ".lower():
            sorted["Adjektive"].append(word)
        elif word["wortart"].lower()  == "ADV".lower():
            sorted["Adverbien"].append(word)
        elif word["wortart"].lower()  == "PRON".lower():
            sorted["Pronomen"].append(word)
        elif word["wortart"].lower()  == "KONJ".lower():
            sorted["Konjunktionen"].append(word)
        elif word["wortart"].lower()  == "PRAEP".lower():
            sorted["Präpositionen"].append(word)
        elif word["wortart"].lower()  == "SUBJ".lower():
            sorted["Subjunktionen"].append(word)
        else:
            sorted["Unbekannt"].append(word)
    return sorted

def split_into_words(input, delete_special_characters=True):
    einzelne_worte = []
    for s in input:
        if delete_special_characters == True:
            s = re.sub(r'[^\w\s]', '', s)
        if len(s.split()) > 1:
            einzelne_worte.extend(s.split())
        else:
            einzelne_worte.append(s)
    return einzelne_worte

def identify_adjectives(word, NomGen, arg, failures, plural):
    for flexion in word["flexion"]:
        if flexion["form"] == f"AdjektivForm(NOM,SG,m,{arg})":
            NomGen["Nom"]["m"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1
        if flexion["form"] == f"AdjektivForm(NOM,SG,f,{arg})":
            NomGen["Nom"]["f"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1
        if flexion["form"] == f"AdjektivForm(NOM,SG,n,{arg})":
            NomGen["Nom"]["n"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1

        if flexion["form"] == f"AdjektivForm(GEN,SG,m,{arg})":
            NomGen["Gen"]["m"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1
        if flexion["form"] == f"AdjektivForm(GEN,SG,f,{arg})":
            NomGen["Gen"]["f"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1
        if flexion["form"] == f"AdjektivForm(GEN,SG,n,{arg})":
            NomGen["Gen"]["n"] = flexion["wort"][0].replace("‑","")
            if not flexion["wort"][0].replace("‑","") == "":
                failures -= 1
    if failures == 6:
        failures = 6
        plural = True
        for flexion in word["flexion"]:
            if flexion["form"] == f"AdjektivForm(NOM,PL,m,{arg})":
                NomGen["Nom"]["m"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1
            if flexion["form"] == f"AdjektivForm(NOM,PL,f,{arg})":
                NomGen["Nom"]["f"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1
            if flexion["form"] == f"AdjektivForm(NOM,PL,n,{arg})":
                NomGen["Nom"]["n"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1

            if flexion["form"] == f"AdjektivForm(GEN,PL,m,{arg})":
                NomGen["Gen"]["m"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1
            if flexion["form"] == f"AdjektivForm(GEN,PL,f,{arg})":
                NomGen["Gen"]["f"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1
            if flexion["form"] == f"AdjektivForm(GEN,PL,n,{arg})":
                NomGen["Gen"]["n"] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0] == "":
                    failures -= 1
    return NomGen, failures, plural

def advanced_formating(input):
    vocabulary = {
        "Nomen": [],#
        "Verben": [],#
        "Adjektive": [],#
        "Adverbien": [],#
        "Pronomen": [],#
        "Konjunktionen": [],#
        "Präpositionen": [],#
        "Subjunktionen": [],#
        "Unbekannt": []#
    }
    for word in input["Nomen"]:
        word_properties = {
            "Nom. Sg.": "-",
            "Gen. Sg.": "-",
            "Genus": "-",
            "Dekl.-Kl.": "-",
            "Bedeutung": "-"
        }
        failures = 2
        for flexion in word["flexion"]:
            if flexion["form"] == "SubstantivForm(NOM,SG)":
                word_properties["Nom. Sg."] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0].replace("‑","") == "":
                    failures -= 1
            if flexion["form"] == "SubstantivForm(GEN,SG)":
                word_properties["Gen. Sg."] = flexion["wort"][0].replace("‑","")
                if not flexion["wort"][0].replace("‑","") == "":
                    failures -= 1
        if failures == 2:
            failures = 2
            for flexion in word["flexion"]:
                if flexion["form"] == "SubstantivForm(NOM,PL)":
                    word_properties["Nom. Sg."] = flexion["wort"][0].replace("‑","") + " (Pl.)"
                    if not flexion["wort"][0].replace("‑","") == "":
                        failures -= 1
                if flexion["form"] == "SubstantivForm(GEN,PL)":
                    word_properties["Gen. Sg."] = flexion["wort"][0].replace("‑","") + " (Pl.)"
                    if not flexion["wort"][0].replace("‑","") == "":
                        failures -= 1
        
        einzelne_worte = split_into_words([word["grundform"].replace("-", " - ")], delete_special_characters=False)
        
        if failures >= 2:
            dice("Fehler: Keine Nominativ- oder Genitivform gefunden", einzelne_worte[0], "bad")

        word_properties["Genus"] = einzelne_worte[-1]+"."
        try:
            if "Dritte Deklination".lower() in word["klassenlabel"].lower():
                word_properties["Dekl.-Kl."] = "kons.-Dekl."
                for flexion in word["flexion"]:
                    if flexion["form"] == "SubstantivForm(GEN,PL)":
                        if flexion["wort"][0].endswith("ium"):
                            word_properties["Dekl.-Kl."] = "gem.-Dekl."
            else:
                word_properties["Dekl.-Kl."] = word["klassenlabel"].replace(",", "").replace("(", "").replace(")", "")[:-7] + "."
        except:
            word_properties["Dekl.-Kl."] = "unbekannt"
        word_properties["Bedeutung"] = word["bedeutungen"]

        vocabulary["Nomen"].append(word_properties)

    for word in input["Verben"]:
        word_properties = {
            "Infinitiv": "-",
            "1. Ps. Sg. Präs. Ind. Akt.": "-",
            "1. Ps. Sg. Perf. Ind. Akt.": "-",
            "PPP": "-",
            "Konj.": "-",
            "Bedeutung": "-"
        }

        "InfinitivForm(PRAES,AKT)"
        "VerbForm(P1,SG,PRAES,IND,AKT,m)"
        "VerbForm(P1,SG,PERF,IND,AKT,m)"
        "PartizipialForm(PPP,AdjektivForm(NOM,SG,n,POS))"

        "VerbForm(P1,SG,PRAES,IND,DEP,m)"
        if word["deponens"] == False:
            for flexion in word["flexion"]:
                if flexion["form"] == "InfinitivForm(PRAES,AKT)":
                    word_properties["Infinitiv"] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "VerbForm(P1,SG,PRAES,IND,AKT,m)":
                    word_properties["1. Ps. Sg. Präs. Ind. Akt."] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "VerbForm(P1,SG,PERF,IND,AKT,m)":
                    word_properties["1. Ps. Sg. Perf. Ind. Akt."] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "PartizipialForm(PPP,AdjektivForm(NOM,SG,n,POS))":
                    word_properties["PPP"] = flexion["wort"][0].replace("‑","")
        else:
            for flexion in word["flexion"]:
                if flexion["form"] == "InfinitivForm(PRAES,DEP)":
                    word_properties["Infinitiv"] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "VerbForm(P1,SG,PRAES,IND,DEP,m)":
                    word_properties["1. Ps. Sg. Präs. Ind. Akt."] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "VerbForm(P1,SG,PERF,IND,DEP,m)":
                    word_properties["1. Ps. Sg. Perf. Ind. Akt."] = flexion["wort"][0].replace("‑","")
                if flexion["form"] == "PartizipialForm(PPP,AdjektivForm(NOM,SG,n,POS))":
                    word_properties["PPP"] = flexion["wort"][0].replace("‑","")

        if "Verb".lower() in word["klassenlabel"].lower():
            word_properties["Konj."] = "unbekannt"
        elif word["deponens"] == True:
            word_properties["Konj."] = "Deponens"
        else:
            word_properties["Konj."] = word["klassenlabel"].replace("ugation", ".")

        word_properties["Bedeutung"] = word["bedeutungen"]

        vocabulary["Verben"].append(word_properties)

    for word in input["Adjektive"]:
        word_properties = {
            "Nom. Sg.: m./f./n.": "",
            "Gen. Sg.: m./f./n.": "",
            "Dekl.-Kl.": "-",
            "Bedeutung": "-"
        }

        failures = 6
        NomGen = {"Nom": {"m": "-", "f": "-", "n": "-"}, "Gen": {"m": "-", "f": "-", "n": "-"}}
        arg = "POS"
        plural = False
        NomGen, failures, plural = identify_adjectives(word, NomGen, arg, failures, plural)
        try:
            if "Adjektiv".lower() in word["klassenlabel"].lower():
                word_properties["Dekl.-Kl."] = "unbekannt"
            elif "Dritte Deklination".lower() in word["klassenlabel"].lower():
                word_properties["Dekl.-Kl."] = "kons.-Dekl."
                for flexion in word["flexion"]:
                    if flexion["form"] == "AdjektivForm(GEN,PL,m,POS)":
                        if flexion["wort"][0].endswith("ium"):
                            word_properties["Dekl.-Kl."] = "gem.-Dekl."
            else:
                word_properties["Dekl.-Kl."] = word["klassenlabel"].replace(",", "").replace("(", "").replace(")", "")[:-7] + "."
        except:
            word_properties["Dekl.-Kl."] = "unbekannt"
        if failures == 6:
            failures = 6
            arg = "KOMP"
            plural = False
            NomGen, failures, plural = identify_adjectives(word, NomGen, arg, failures, plural)
            word_properties["Dekl.-Kl."] = "Komp."
            if failures == 6:
                failures = 6
                arg = "SUP"
                plural = False
                NomGen, failures, plural = identify_adjectives(word, NomGen, arg, failures, plural)
                word_properties["Dekl.-Kl."] = "Sup."
                if failures == 6:
                    failures = 6
                    plural = False
                    word_properties["Dekl.-Kl."] = "unbekannt"
                    dice("Es konnten keine der Flexionen extrahiert werden", einzelne_worte[0], "bad")

        einzelne_worte = split_into_words([word["grundform"].replace("-", " - ")], delete_special_characters=False)
        for key in NomGen:
            NomGen[key] = ', '.join(NomGen[key].values())
        if plural == True:
            word_properties["Nom. Sg.: m./f./n."] = NomGen["Nom"] + " (Pl.)"
            word_properties["Gen. Sg.: m./f./n."] = NomGen["Gen"] + " (Pl.)"
        else:
            word_properties["Nom. Sg.: m./f./n."] = NomGen["Nom"]
            word_properties["Gen. Sg.: m./f./n."] = NomGen["Gen"]

        word_properties["Bedeutung"] = word["bedeutungen"]

        vocabulary["Adjektive"].append(word_properties)

    return vocabulary

def save2json(file, data):
    with open(file,"w+",encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

def save2excel(filepath, vocabulary, cleanup=True):
    if cleanup == True:
        tmp_excel = BytesIO()
        with pd.ExcelWriter(tmp_excel) as writer:
            # Für jede Wortart ein eigenes Sheet erstellen
            for wortart, vokabeln in vocabulary.items():
                # Ein DataFrame für jede Wortart erstellen
                df = pd.DataFrame(vokabeln)
                # In das entsprechende Sheet schreiben
                df.to_excel(writer, sheet_name=wortart, index=False)
        file = cleanup_excel(tmp_excel, filepath)
        return file
    elif cleanup == False:
        with pd.ExcelWriter(filepath) as writer:
            # Für jede Wortart ein eigenes Sheet erstellen
            for wortart, vokabeln in vocabulary.items():
                # Ein DataFrame für jede Wortart erstellen
                df = pd.DataFrame(vokabeln)
                # In das entsprechende Sheet schreiben
                df.to_excel(writer, sheet_name=wortart, index=False)

def locale_sort_key(value):
    return locale.strxfrm(value)

def cleanup_excel(input_file, output_file):
    excel_file = pd.ExcelFile(input_file)

    output_stream = BytesIO()

    sheets_dict = {}

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        if df.empty:
            continue
        
        df_no_duplicates = df.drop_duplicates()

        if df_no_duplicates.empty:
            continue
        
        #df_sorted = df_no_duplicates.sort_values(by=df_no_duplicates.columns[0], ascending=True)
        df_sorted = df_no_duplicates.sort_values(by=df_no_duplicates.columns[0], key=lambda col: col.map(locale_sort_key), ascending=True)
        sheets_dict[sheet_name] = df_sorted

    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_stream)
    
    # Definiere den grauen Farbton und die Rahmeneinstellungen
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Anpassung der Spaltenbreite und Stile für Spaltennamen und Zellen
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
       
        # Grau für die Spaltennamen und Rahmenlinien für alle Zellen
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border  # Rahmenlinien für jede Zelle
                if cell.row == 1:  # Überprüfen, ob es sich um die Kopfzeile handelt
                    cell.fill = gray_fill  # Grau für die Kopfzeile

        # Dynamische Anpassung der Spaltenbreite
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Holt den Spaltenbuchstaben
            for cell in column_cells:
                try:
                    if cell.value:  # Nur nicht-leere Zellen prüfen
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
        
        ws = sheet

        # Bestimmen der letzten Zeile und der Anzahl der Spalten
        letzte_zeile = ws.max_row
        letzte_spalte = ws.max_column

        # Die Zellen am Ende des Dokuments zusammenführen
        ws.merge_cells(start_row=letzte_zeile + 1, start_column=1, end_row=letzte_zeile + 1, end_column=letzte_spalte)

        kursiv_font = Font(italic=True)
        # Text in die gemergte Zelle einfügen
        zelle = ws.cell(row=letzte_zeile + 1, column=1, value="Diese Liste basiert auf Daten von \"https://www.navigium.de/suchfunktion/_search?q={}\". © Rechteinhaber: Navigium.de")
        zelle.font = kursiv_font

    wb.save(output_file)
    wb.save(input_file)
    print(f"Bereinigte Datei wurde als '{output_file}' gespeichert, und die Spaltenbreiten wurden angepasst.")
    return input_file

def save2by2(excel_file, by2_filepath):
    Data = {}

    # Lade die Excel-Datei und hole die Namen der Arbeitsblätter
    excel_file = pd.ExcelFile(excel_file)
    sheet_names = ["Nomen", "Verben", "Adjektive", "Adverbien", "Pronomen", "Konjunktionen", "Subjunktionen", "Präpositionen", "Unbekannt"]

    # Iteriere über jedes Arbeitsblatt in der Excel-Datei
    for sheet in sheet_names:
        # Lese das Arbeitsblatt in ein DataFrame
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet)
        except:
            continue
        # Hole Fragen und Antworten aus den ersten beiden Spalten
        fragen = df.iloc[:, 0].tolist()
        if sheet == "Nomen" or sheet == "Verben" or sheet == "Adjektive" or sheet == "Pronomen" or sheet == "Präpositionen":
            result_list = []
            for index, row in df.iterrows():
                # Greife auf die zweite und dritte Spalte der Zeile zu, unabhängig von den Namen
                if sheet == "Nomen":
                    combined_str = f"{row.iloc[0]}, {row.iloc[1]}; {row.iloc[2]}; {row.iloc[3]}:\n{row.iloc[-1]}"
                elif sheet == "Verben":
                    combined_str = f"{row.iloc[0]}, {row.iloc[1]}, {row.iloc[2]}, {row.iloc[3]}; {row.iloc[4]}:\n{row.iloc[-1]}"
                elif sheet == "Adjektive":
                    combined_str = f"{row.iloc[0]}; {row.iloc[1]}; {row.iloc[2]}:\n{row.iloc[-1]}"
                elif sheet == "Präpositionen":
                    combined_str = f"{row.iloc[0]} ({row.iloc[1]}):\n{row.iloc[-1]}"
                elif sheet == "Pronomen":
                    combined_str = f"{row.iloc[0]}, {row.iloc[1]}:\n{row.iloc[-1]}"                  
                # Füge den kombinierten String der Liste hinzu
                result_list.append(combined_str)
            antworten = result_list
        else:
            antworten = df.iloc[:, 1].tolist()

        # Speichere die Fragen und Antworten im Dictionary
        Data[sheet] = (fragen, antworten)
    save_by2(by2_filepath, Data)